from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OPEN_COLUMNS = [
    ("FOR ID", 10),
    ("Context Summary", 42),
    ("Requester", 16),
    ("Current Owner", 22),
    ("Action Items", 42),
    ("Priority", 12),
    ("Status", 26),
    ("Carry-Over", 13),
    ("Due Date", 14),
    ("Confidence", 14),
    ("Confidence Notes", 38),
    ("Flags", 28),
    ("Flag Notes", 38),
]

CARRY_OVER_COLORS = {
    3: "FFF2CC",  # yellow
    4: "FFD700",  # amber
}
CARRY_OVER_COLOR_5PLUS = "FF6B6B"  # red

CLOSED_COLUMNS = [
    ("FOR ID", 10),
    ("Context Summary", 42),
    ("Requester", 16),
    ("Resolved By", 20),
    ("Resolution", 42),
    ("Date Closed", 16),
]

WATCH_COLUMNS = [
    ("Watch ID", 10),
    ("Description", 50),
    ("Raised By", 18),
    ("Next Step", 38),
    ("Notes", 42),
    ("Meeting Date", 16),
]

PRIORITY_COLORS = {
    "High":   "FFCCCC",
    "Medium": "FFF2CC",
    "Low":    "CCFFCC",
}

CONFIDENCE_COLORS = {
    "Low":    "FF6B6B",
    "Medium": "FFD93D",
    "High":   "CCFFCC",
}

HEADER_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_GREEN = PatternFill("solid", fgColor="375623")
HEADER_FILL_GRAY  = PatternFill("solid", fgColor="404040")
HEADER_FILL_ORANGE = PatternFill("solid", fgColor="7F4E1F")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
FLAG_FILL   = PatternFill("solid", fgColor="FFE0B2")
LOW_CONF_ROW_TINT = "FFF0F0"

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_header(ws, columns, fill):
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28


def _confidence_label(owner_conf: str, priority_conf: str) -> str:
    parts = []
    if owner_conf != "High":
        parts.append(f"Owner: {owner_conf}")
    if priority_conf != "High":
        parts.append(f"Priority: {priority_conf}")
    return ", ".join(parts) if parts else "High"


def _confidence_fill(owner_conf: str, priority_conf: str) -> PatternFill:
    if owner_conf == "Low" or priority_conf == "Low":
        return PatternFill("solid", fgColor=CONFIDENCE_COLORS["Low"])
    if owner_conf == "Medium" or priority_conf == "Medium":
        return PatternFill("solid", fgColor=CONFIDENCE_COLORS["Medium"])
    return PatternFill("solid", fgColor=CONFIDENCE_COLORS["High"])


def _open_items_sheet(ws, action_items: list[dict]):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(OPEN_COLUMNS))}1"
    _make_header(ws, OPEN_COLUMNS, HEADER_FILL_BLUE)

    for i, item in enumerate(action_items, start=2):
        flags = item.get("flags") or []
        flag_str = ", ".join(flags) if flags else ""
        priority = item.get("priority", "Medium")
        owner_conf = item.get("owner_confidence", "High")
        priority_conf = item.get("priority_confidence", "High")
        has_low_conf = owner_conf == "Low" or priority_conf == "Low"
        consecutive = item.get("consecutive_meetings", 1)
        is_carry_over = item.get("carry_over_flag", False)

        conf_label = _confidence_label(owner_conf, priority_conf)
        conf_fill = _confidence_fill(owner_conf, priority_conf)

        if is_carry_over:
            carry_label = f"{consecutive} meetings"
            carry_color = CARRY_OVER_COLOR_5PLUS if consecutive >= 5 else CARRY_OVER_COLORS.get(consecutive, CARRY_OVER_COLOR_5PLUS)
            carry_fill = PatternFill("solid", fgColor=carry_color)
        else:
            carry_label = ""
            carry_fill = None

        row_base_fill = PatternFill("solid", fgColor=LOW_CONF_ROW_TINT if has_low_conf else ("F9F9F9" if i % 2 == 0 else "FFFFFF"))

        values = [
            item.get("request_id", i - 1),
            item.get("context_summary", ""),
            item.get("requester", ""),
            item.get("current_owner", ""),
            item.get("action_items", ""),
            priority,
            item.get("status", ""),
            carry_label,
            item.get("due_date", "Not specified"),
            conf_label,
            item.get("confidence_notes", ""),
            flag_str,
            item.get("flag_notes", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx == 6:  # Priority
                cell.fill = PatternFill("solid", fgColor=PRIORITY_COLORS.get(priority, "FFFFFF"))
                cell.font = Font(bold=True, name="Calibri", size=10)
            elif col_idx == 8:  # Carry-Over
                cell.fill = carry_fill if carry_fill else row_base_fill
                cell.font = Font(bold=is_carry_over, name="Calibri", size=10)
            elif col_idx == 10:  # Confidence
                cell.fill = conf_fill
                cell.font = Font(bold=has_low_conf, name="Calibri", size=10)
            elif flag_str and col_idx == 12:  # Flags
                cell.fill = FLAG_FILL
            else:
                cell.fill = row_base_fill

        ws.row_dimensions[i].height = 55

    summary_row = len(action_items) + 2
    ws.cell(row=summary_row, column=1, value=f"Total open: {len(action_items)}").font = Font(bold=True, name="Calibri", size=10)

    carry_count = sum(1 for i in action_items if i.get("carry_over_flag"))
    if carry_count:
        cc = ws.cell(row=summary_row, column=8, value=f"Carry-overs: {carry_count}")
        cc.font = Font(bold=True, name="Calibri", size=10)
        cc.fill = PatternFill("solid", fgColor=CARRY_OVER_COLORS[3])

    low_conf_count = sum(1 for i in action_items if i.get("owner_confidence") == "Low" or i.get("priority_confidence") == "Low")
    if low_conf_count:
        lc = ws.cell(row=summary_row, column=10, value=f"Low confidence: {low_conf_count}")
        lc.font = Font(bold=True, name="Calibri", size=10)
        lc.fill = PatternFill("solid", fgColor=CONFIDENCE_COLORS["Low"])

    flagged = sum(1 for i in action_items if i.get("flags"))
    if flagged:
        fc = ws.cell(row=summary_row, column=12, value=f"Flagged: {flagged}")
        fc.font = Font(bold=True, name="Calibri", size=10)
        fc.fill = FLAG_FILL

    ws.cell(
        row=summary_row + 1, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ).font = Font(italic=True, color="888888", name="Calibri", size=9)


def _closed_items_sheet(ws, closed_items: list[dict]):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CLOSED_COLUMNS))}1"
    _make_header(ws, CLOSED_COLUMNS, HEADER_FILL_GREEN)

    for i, item in enumerate(closed_items, start=2):
        values = [
            item.get("request_id", ""),
            item.get("context_summary", ""),
            item.get("requester", ""),
            item.get("resolved_by", ""),
            item.get("resolution", ""),
            item.get("date_closed", "Unknown"),
        ]
        row_fill = PatternFill("solid", fgColor="EEF4EE" if i % 2 == 0 else "FFFFFF")
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill
        ws.row_dimensions[i].height = 50

    summary_row = len(closed_items) + 2
    ws.cell(row=summary_row, column=1, value=f"Total closed (all time): {len(closed_items)}").font = Font(bold=True, name="Calibri", size=10)


def _watch_items_sheet(ws, watch_items: list[dict]):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(WATCH_COLUMNS))}1"
    _make_header(ws, WATCH_COLUMNS, HEADER_FILL_ORANGE)

    for i, item in enumerate(watch_items, start=2):
        values = [
            item.get("item_id", i - 1),
            item.get("description", ""),
            item.get("raised_by", "Unknown"),
            item.get("next_step", "Monitor"),
            item.get("notes", ""),
            item.get("meeting_date", "Unknown"),
        ]
        row_fill = PatternFill("solid", fgColor="FFF8F0" if i % 2 == 0 else "FFFFFF")
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = row_fill
        ws.row_dimensions[i].height = 50

    summary_row = len(watch_items) + 2
    ws.cell(row=summary_row, column=1, value=f"Total watch items (all time): {len(watch_items)}").font = Font(bold=True, name="Calibri", size=10)


def _summary_sheet(ws, meeting_history: list[dict]):
    _make_header(ws, [("Meeting History — Latest First", 80)], HEADER_FILL_GRAY)

    current_row = 2
    for idx, meeting in enumerate(meeting_history):
        is_latest = idx == 0
        banner_fill = PatternFill("solid", fgColor="2E4057" if is_latest else "666666")
        banner = ws.cell(
            row=current_row, column=1,
            value=f"{'LATEST — ' if is_latest else ''}Meeting: {meeting.get('date', 'Unknown')}  |  Generated: {meeting.get('generated_at', 'Unknown')}"
        )
        banner.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        banner.fill = banner_fill
        banner.alignment = Alignment(vertical="center", wrap_text=False)
        banner.border = BORDER
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for label, value in [
            ("Duration", meeting.get("duration", "Unknown")),
            ("Attendees", ", ".join(meeting.get("attendees", []))),
            ("Summary", meeting.get("summary", "")),
        ]:
            cell = ws.cell(row=current_row, column=1, value=f"{label}:  {value}")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
            if label == "Summary":
                cell.font = Font(name="Calibri", size=11)
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
                ws.row_dimensions[current_row].height = 90
            else:
                cell.font = Font(name="Calibri", size=11)
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
                ws.row_dimensions[current_row].height = 20
            current_row += 1

        ws.row_dimensions[current_row].height = 8
        current_row += 1

    ws.column_dimensions["A"].width = 100


def save_excel(result: dict, closed_history: list[dict], watch_history: list[dict], meeting_history: list[dict], output_path: str):
    meeting = result.get("meeting", {})
    open_items = result.get("open_items", [])
    new_closed = result.get("closed_items", [])
    new_watch = result.get("watch_items", [])

    # Merge closed items (dedupe by request_id)
    existing_closed_ids = {str(c.get("request_id", "")).strip() for c in closed_history}
    for item in new_closed:
        if str(item.get("request_id", "")).strip() not in existing_closed_ids:
            closed_history.append(item)

    # Merge watch items (dedupe by description similarity via item_id)
    existing_watch_ids = {str(w.get("item_id", "")).strip() for w in watch_history}
    for item in new_watch:
        item["meeting_date"] = meeting.get("date", "Unknown")
        if str(item.get("item_id", "")).strip() not in existing_watch_ids:
            watch_history.append(item)

    wb = Workbook()

    ws_open = wb.active
    ws_open.title = "Open Action Items"
    _open_items_sheet(ws_open, open_items)

    ws_closed = wb.create_sheet("Closed Items")
    _closed_items_sheet(ws_closed, closed_history)

    ws_watch = wb.create_sheet("Watch Items")
    _watch_items_sheet(ws_watch, watch_history)

    ws_summary = wb.create_sheet("Meeting Summary")
    _summary_sheet(ws_summary, meeting_history)

    wb.save(output_path)
    print(f"Excel saved: {output_path}")

    return closed_history, watch_history
